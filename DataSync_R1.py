import glob
import warnings
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import numbers
from collections import defaultdict
import pandas as pd

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

def extract_rut_numbers(rut):
    return ''.join(filter(str.isdigit, str(rut)))

def get_real_protocolo(rut_num, patients_df, fecha_corte):
    patient_records = patients_df[
        (patients_df['RUT_NUM'] == rut_num) & (patients_df['FECHA_FIRMA_CI'] <= fecha_corte)
    ]
    if not patient_records.empty:
        most_recent = patient_records.sort_values('FECHA_FIRMA_CI', ascending=False).iloc[0]
        return most_recent['Protocolo']
    return 'SIN_PROTOCOLO'

def actualizar_tipo_prestacion(row, ajuste_prestacion_map):
    if row['ORIGEN'] == 'CAJA':
        return ajuste_prestacion_map.get(row['PRESTACION'], row['TIPO_PRESTACION'])
    return row['TIPO_PRESTACION']

def process_sheet(sheet_name, budget_file, output_path, base_path, eur_col, usd_col, clp_col):
    external_file_path = os.path.join(base_path, budget_file)
    try:
        external_wb = load_workbook(external_file_path, read_only=True, data_only=True)
        external_ws = external_wb.active
        budget_values = {}
        for row in external_ws.iter_rows(min_row=2, values_only=True):
            key = str(row[0]).strip()
            budget_values[key] = {'EUR': row[eur_col], 'USD': row[usd_col], 'CLP': row[clp_col]}
    except Exception as e:
        print(f"Error procesando {external_file_path}: {e}")
        budget_values = {}

    wb = load_workbook(output_path, keep_links=True)
    ws = wb[sheet_name]

    # Identificar columnas
    col_eur = col_usd = col_clp = col_total = col_utilidad = None
    for col in ws.iter_cols(1, ws.max_column):
        header = col[0].value
        if header == 'EUR': col_eur = col[0].column_letter
        elif header == 'USD': col_usd = col[0].column_letter
        elif header == 'CLP': col_clp = col[0].column_letter
        elif header == 'TOTAL': col_total = col[0].column_letter
        elif header == 'UTILIDAD': col_utilidad = col[0].column_letter

    # Procesar filas
    for row in range(2, ws.max_row + 1):
        key = str(ws[f'A{row}'].value).strip()
        if key in budget_values:
            eur_val = budget_values[key]['EUR']
            usd_val = budget_values[key]['USD']
            clp_val = budget_values[key]['CLP']

            if col_eur and eur_val is not None: ws[f'{col_eur}{row}'] = eur_val
            if col_usd and usd_val is not None: ws[f'{col_usd}{row}'] = usd_val
            if col_clp and clp_val is not None: ws[f'{col_clp}{row}'] = round(clp_val)

            if col_clp and col_total and col_utilidad:
                total_val = ws[f'{col_total}{row}'].value
                if clp_val is not None and total_val is not None:
                    try:
                        utilidad = float(clp_val) - float(total_val)
                        ws[f'{col_utilidad}{row}'] = round(utilidad)
                    except ValueError: pass

    # Agregar SUBTOTALES para PDC_LAB y PDF
    subtotal_clp = None
    if sheet_name in ["PDC_LAB", "PDF"] and col_total and col_clp and col_utilidad:
        total_sum = clp_sum = utilidad_sum = 0
        for row in range(2, ws.max_row + 1):
            total_val = ws[f'{col_total}{row}'].value or 0
            clp_val = ws[f'{col_clp}{row}'].value or 0
            utilidad_val = ws[f'{col_utilidad}{row}'].value or 0

            total_sum += total_val
            clp_sum += clp_val
            utilidad_sum += utilidad_val

        ws.append([])
        subtotal_row = ws.max_row + 1
        ws[f'A{subtotal_row}'] = "SUBTOTAL"
        ws[f'{col_total}{subtotal_row}'] = round(total_sum)
        ws[f'{col_clp}{subtotal_row}'] = round(clp_sum)
        ws[f'{col_utilidad}{subtotal_row}'] = round(utilidad_sum)
        subtotal_clp = round(clp_sum)
    wb.save(output_path)
    wb.close()
    return subtotal_clp

def count_empty_cells(ws, col_letter):
    """Cuenta las celdas vacías en una columna específica de una hoja de trabajo."""
    empty_count = 0
    for row in range(2, ws.max_row + 1):
        if ws[f'{col_letter}{row}'].value is None:
            empty_count += 1
    return empty_count

def add_rentabilidad_sheet(wb, mes, pdf_df, pdc_img_df, pdc_lab_df, subtotal_pdc_lab, subtotal_pdf):
    ws = wb.create_sheet(title="8-Rentabilidad de la Unidad")

    # Calcular los totales para Laboratorio y Quimio
    total_lab = pdc_lab_df['TOTAL'].sum()
    total_quimio = pdf_df['TOTAL'].sum()

    # Filtrar para Prestaciones de Imágenes
    filtro_imagen = ~(pdc_img_df['PRESTACION'].astype(str).str.startswith("E.C.G.") | 
                      pdc_img_df['PRESTACION'].astype(str).str.startswith("Ecocardiograma"))
    total_imagen = pdc_img_df.loc[filtro_imagen, 'TOTAL'].sum()

    # Para ECG + Ecocardio: sumar solo las filas cuya PRESTACION inicie con "E.C.G." o "Ecocardiograma"
    filtro_ecg = pdc_img_df['PRESTACION'].astype(str).str.startswith("E.C.G.") | \
                 pdc_img_df['PRESTACION'].astype(str).str.startswith("Ecocardiograma")
    total_ecg = pdc_img_df.loc[filtro_ecg, 'TOTAL'].sum()

    # Calcular BUDGET (suma de CLP) y UTILIDAD (BUDGET - COSTO)
    budget_lab = pdc_lab_df['CLP'].sum()
    budget_quimio = pdf_df['CLP'].sum()
    budget_imagen = pdc_img_df.loc[filtro_imagen, 'CLP'].sum()
    budget_ecg = pdc_img_df.loc[filtro_ecg, 'CLP'].sum()

    utilidad_lab = budget_lab - total_lab
    utilidad_imagen = budget_imagen - total_imagen
    utilidad_quimio = budget_quimio - total_quimio
    utilidad_ecg = budget_ecg - total_ecg

    data = [
        [mes, "COSTO", "BUDGET", "UTILIDAD", "CHECK"],
        ["Prestaciones de Laboratorio", round(total_lab), round(budget_lab), round(utilidad_lab), ""],
        ["Prestaciones de Imágenes", round(total_imagen), round(budget_imagen), round(utilidad_imagen), ""],
        ["Quimio", round(total_quimio), round(budget_quimio), round(utilidad_quimio), ""],
        ["ECG + Ecocardio", round(total_ecg), round(budget_ecg), round(utilidad_ecg), ""],
        ["Coordinacion", "", "", "", ""],
        ["OH", "", "", "", ""],
        ["Honorarios Médicos", "", "", "", ""],
        ["Otros Ingresos", "", "", "", ""],
        ["Remuneraciones", "", "", "", ""],
        ["Transporte pacientes", "", "", "", ""],
        ["Costos preparaciones oncologicas", "", "", "", ""],
        ["Otros costos y gastos", "", "", "", ""]
    ]
    for row in data:
        ws.append(row)
    
    # Trasladar el valor del subtotal de la hoja PDC_LAB a C2 de la hoja 8-Rentabilidad de la Unidad
    if subtotal_pdc_lab is not None:
        ws['C2'] = subtotal_pdc_lab
    
    # Calcular D2 como la resta entre C2 y B2
    ws['D2'] = ws['C2'].value - ws['B2'].value

    # Trasladar el valor del subtotal de la hoja PDF a C4 de la hoja 8-Rentabilidad de la Unidad
    if subtotal_pdf is not None:
        ws['C4'] = subtotal_pdf
    
    # Calcular D4 como la resta entre C4 y B4
    ws['D4'] = ws['C4'].value - ws['B4'].value

    # Contar celdas vacías en las hojas especificadas y llenar la columna "CHECK"
    if 'PDC_LAB' in wb.sheetnames:
        ws_pdc_lab = wb['PDC_LAB']
        ws['E2'] = count_empty_cells(ws_pdc_lab, 'S')  # Contar celdas vacías en la columna S de PDC_LAB

    if 'PDC_IMG_NONECG' in wb.sheetnames:
        ws_pdc_img_nonecg = wb['PDC_IMG_NONECG']
        ws['E3'] = count_empty_cells(ws_pdc_img_nonecg, 'S')  # Contar celdas vacías en la columna S de PDC_IMG_NONECG

    if 'PDF' in wb.sheetnames:
        ws_pdf = wb['PDF']
        ws['E4'] = count_empty_cells(ws_pdf, 'S')  # Contar celdas vacías en la columna S de PDF

    if 'PDC_IMG_ECG' in wb.sheetnames:
        ws_pdc_img_ecg = wb['PDC_IMG_ECG']
        ws['E5'] = count_empty_cells(ws_pdc_img_ecg, 'S')  # Contar celdas vacías en la columna S de PDC_IMG_ECG

    # Aplicar formato de números sin decimales
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=4):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0'

    # Ajustar el ancho de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Ajustar el ancho de la columna A (descripción) para que sea más ancha
    ws.column_dimensions['A'].width = 30

    # Mover la hoja "8-Rentabilidad de la Unidad" al final
    wb._sheets = [sheet for sheet in wb._sheets if sheet.title != "8-Rentabilidad de la Unidad"] + [ws]

def main():
    """
    Main function to generate a monthly report based on protocol and patient data.
    
    This function performs the following steps:
    1. Prompts the user to input a month and year for filtering.
    2. Loads various Excel files containing protocol, patient, and dictionary data.
    3. Processes the data to generate a consolidated report.
    4. Saves the report to an output Excel file.
    
    Usage:
    Run the script and follow the prompt to input the month and year in MMYY format.
    Ensure that the required Excel files are present in the specified base path.
    """
    base_path = r'C:\Users\jorge.castros\Documents\Reporte Mensual'
    fecha_filtro = input("Ingrese el mes y año para filtrar (formato MMYY): ")
    mes_filtro = int(fecha_filtro[:2])
    anio_filtro = int("20" + fecha_filtro[2:])
    fecha_corte = pd.Timestamp(year=anio_filtro, month=mes_filtro, day=1) + pd.offsets.MonthEnd(0)

    # Diccionario para traducir los nombres de los meses al español
    meses_espanol = {
        'January': 'ENERO', 'February': 'FEBRERO', 'March': 'MARZO', 'April': 'ABRIL',
        'May': 'MAYO', 'June': 'JUNIO', 'July': 'JULIO', 'August': 'AGOSTO',
        'September': 'SEPTIEMBRE', 'October': 'OCTUBRE', 'November': 'NOVIEMBRE', 'December': 'DICIEMBRE'
    }
    mes_nombre = meses_espanol[fecha_corte.strftime('%B')]

    output_folder = os.path.join(base_path, f'{fecha_filtro[:2]}')
    os.makedirs(output_folder, exist_ok=True)

    protocol_files = glob.glob(os.path.join(base_path, "Protocolo*.xlsx"))
    if not protocol_files:
        print("No se encontró ningún archivo que comience con 'Protocolo' en la carpeta especificada.")
        return

    protocol_path = protocol_files[0]  # Seleccionar el primer archivo de la lista
    protocol_df = pd.read_excel(protocol_path, engine='openpyxl')

    patients_path = os.path.join(base_path, 'Base de Datos Pacientes UIDO.xlsx')
    patients_df = pd.read_excel(patients_path, sheet_name='Consolidado', engine='openpyxl')

    diccionario_path = os.path.join(base_path, 'Diccionario_Protocolos.xlsx')
    diccionario_df = pd.read_excel(diccionario_path, sheet_name='AJUSTE_PRESTACION', engine='openpyxl')
    dicc_prot_df = pd.read_excel(diccionario_path, sheet_name='DICC_PROT', engine='openpyxl')

    required_columns_protocol = ['RUT_PACIENTE', 'CONVENIO', 'FECHA', 'PRESTACION', 'TIPO_PRESTACION', 'ORIGEN', 'MOTIVOINGRESO', 'folio', 'EXENTO', 'AFECTO', 'IVA', 'NETO', 'TOTAL']
    required_columns_patients = ['RUT (sin puntos y con guion)', 'Fecha Firma CI (dd-mm-aaaa)', 'Protocolo']
    required_columns_diccionario = ['PRESTACION', 'TIPO_PRESTACION', 'ORIGEN']
    required_columns_dicc_prot = ['Protocolo_Pacientes', 'CONVENIO']

    if not all(col in protocol_df.columns for col in required_columns_protocol):
        print(f"Faltan columnas en el archivo de protocolos: {required_columns_protocol}")
        return
    if not all(col in patients_df.columns for col in required_columns_patients):
        print(f"Faltan columnas en el archivo de pacientes: {required_columns_patients}")
        return
    if not all(col in diccionario_df.columns for col in required_columns_diccionario):
        print(f"Faltan columnas en el archivo del diccionario: {required_columns_diccionario}")
        return
    if not all(col in dicc_prot_df.columns for col in required_columns_dicc_prot):
        print(f"Faltan columnas en la hoja DICC_PROT del diccionario: {required_columns_dicc_prot}")
        return

    protocol_df['RUT_NUM'] = protocol_df['RUT_PACIENTE'].apply(extract_rut_numbers)
    patients_df['RUT_NUM'] = patients_df['RUT (sin puntos y con guion)'].apply(extract_rut_numbers)

    protocol_df['FECHA'] = pd.to_datetime(protocol_df['FECHA'], errors='coerce')
    patients_df['FECHA_FIRMA_CI'] = pd.to_datetime(patients_df['Fecha Firma CI (dd-mm-aaaa)'], errors='coerce')

    protocol_df = protocol_df[
        (protocol_df['FECHA'].dt.month == mes_filtro) &
        (protocol_df['FECHA'].dt.year == anio_filtro)
    ]

    rut_convenio_counts = protocol_df.groupby('RUT_NUM')['CONVENIO'].nunique()
    protocol_df['ES_DUP'] = protocol_df['RUT_NUM'].apply(
        lambda x: 'Duplicado' if rut_convenio_counts[x] > 1 else 'Unico'
    )

    protocolo_to_convenio = dicc_prot_df.set_index('Protocolo_Pacientes')['CONVENIO'].to_dict()

    protocol_df['Protocolo Real'] = protocol_df.apply(
        lambda row: row['CONVENIO'] if row['ES_DUP'] == 'Unico' else get_real_protocolo(row['RUT_NUM'], patients_df, fecha_corte),
        axis=1
    )

    protocol_df.loc[protocol_df['ES_DUP'] == 'Duplicado', 'Protocolo Real'] = protocol_df.loc[
        protocol_df['ES_DUP'] == 'Duplicado', 'Protocolo Real'
    ].map(protocolo_to_convenio)

    diccionario_df = diccionario_df.drop_duplicates(subset='PRESTACION')

    ajuste_prestacion_df = pd.read_excel(diccionario_path, sheet_name='AJUSTE_PRESTACION', engine='openpyxl')
    ajuste_prestacion_map = ajuste_prestacion_df.set_index('PRESTACION')['TIPO_PRESTACION'].to_dict()

    protocol_df['TIPO_PRESTACION'] = protocol_df.apply(lambda row: actualizar_tipo_prestacion(row, ajuste_prestacion_map), axis=1)

    pdf_df = protocol_df[
        (protocol_df['ORIGEN'] == 'FOLIO') &
        (protocol_df['MOTIVOINGRESO'] == 'QUIMIOTERAPIA') &
        (~protocol_df['TIPO_PRESTACION'].isin(['Farmaco no Oncologico', 'Farmaco Oncologico']))
    ]

    pdf_df_consolidated = pdf_df.groupby('folio').agg({
        'RUT_PACIENTE': 'first',
        'CONVENIO': 'first',
        'FECHA': 'first',
        'PRESTACION': 'first',
        'TIPO_PRESTACION': 'first',
        'ORIGEN': 'first',
        'MOTIVOINGRESO': 'first',
        'EXENTO': 'sum',
        'AFECTO': 'sum',
        'IVA': 'sum',
        'NETO': 'sum',
        'TOTAL': 'sum',
        'Protocolo Real': 'first'
    }).reset_index()

    pdf_df = pdf_df.drop(columns=['EXENTO', 'AFECTO', 'IVA', 'NETO', 'TOTAL']).drop_duplicates(subset=['folio'])
    pdf_df = pdf_df.merge(pdf_df_consolidated[['folio', 'EXENTO', 'AFECTO', 'IVA', 'NETO', 'TOTAL']], on='folio', how='left')
    pdf_df = pdf_df[protocol_df.columns]

    pdc_df = protocol_df[protocol_df['ORIGEN'] == 'CAJA']

    for col in ['UTILIDAD','CLP','USD','EUR']:
        pdf_df.insert(pdf_df.columns.get_loc('TOTAL') + 1, col, None)
        pdc_df.insert(pdc_df.columns.get_loc('TOTAL') + 1, col, None)

    pdc_img_df = pdc_df[pdc_df['TIPO_PRESTACION'].isin(['Imagen', 'Imagenología'])].copy()
    pdc_img_df.loc[:, 'KEY'] = pdc_img_df['CODIGO'].astype(str).str.strip() + pdc_img_df['Protocolo Real'].astype(str).str.strip()
    pdc_img_df.insert(0, 'KEY', pdc_img_df.pop('KEY'))

    pdc_lab_df = pdc_df[pdc_df['TIPO_PRESTACION'] == 'Laboratorio'].copy()
    pdc_lab_df.loc[:, 'KEY'] = pdc_lab_df['CODIGO'].astype(str).str.strip() + pdc_lab_df['Protocolo Real'].astype(str).str.strip()
    pdc_lab_df.insert(0, 'KEY', pdc_lab_df.pop('KEY'))

    pdf_df.loc[:, 'KEY'] = pdf_df['folio'].astype(int).astype(str).str.strip() + pdf_df['Protocolo Real'].astype(str).str.strip()
    pdf_df.insert(0, 'KEY', pdf_df.pop('KEY'))

    output_path = os.path.join(output_folder, f'Monthly_Report_{fecha_filtro}.xlsx')
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        protocol_df.to_excel(writer, sheet_name=f'BBDD_{fecha_filtro}', index=False)
        pdf_df.to_excel(writer, sheet_name='PDF', index=False)
        pdc_df.to_excel(writer, sheet_name='PDC', index=False)
        pdc_img_df.to_excel(writer, sheet_name='PDC_IMG', index=False)
        pdc_lab_df.to_excel(writer, sheet_name='PDC_LAB', index=False)

    # Procesar las hojas PDC_IMG, PDC_LAB y PDF con redondeo al entero más cercano para CLP y UTILIDAD
    process_sheet('PDC_IMG', 'BBDD_BUDGET_IMG.xlsx', output_path, base_path, 8, 9, 10)
    subtotal_pdc_lab = process_sheet('PDC_LAB', 'BBDD_BUDGET_LAB.xlsx', output_path, base_path, 15, 16, 17)
    subtotal_pdf = process_sheet('PDF', 'BBDD_BUDGET_FOLIO.xlsx', output_path, base_path, 17, 18, 19)

    # Verificar si subtotal_pdf es None
    if subtotal_pdf is None:
        print("Advertencia: No se pudo calcular el subtotal para PDF.")
        subtotal_pdf = 0  # O algún valor por defecto

    # Agregar la hoja "8-Rentabilidad de la Unidad"
    wb = load_workbook(output_path)
    add_rentabilidad_sheet(wb, mes_nombre, pdf_df, pdc_img_df, pdc_lab_df, subtotal_pdc_lab, subtotal_pdf)

    # Crear la hoja PDC_IMG_ECG con la información filtrada de PDC_IMG
    if 'PDC_IMG' in wb.sheetnames:
        ws_pdc_img = wb['PDC_IMG']
        ws_pdc_img_ecg = wb.create_sheet(title='PDC_IMG_ECG')

        # Copiar encabezados como una fila
        headers = [cell.value for cell in ws_pdc_img[1]]
        ws_pdc_img_ecg.append(headers)  # Añadir los encabezados como una fila

        # Encontrar el índice de la columna 'PRESTACION'
        prestacion_col_idx = None
        for idx, cell in enumerate(ws_pdc_img[1]):
            if cell.value == 'PRESTACION':
                prestacion_col_idx = idx
                break

        if prestacion_col_idx is None:
            print("No se encontró la columna 'PRESTACION' en la hoja PDC_IMG.")
        else:
            # Filtrar filas y copiar datos
            for row in ws_pdc_img.iter_rows(min_row=2, values_only=True):
                prestacion = str(row[prestacion_col_idx])
                if prestacion.startswith("E.C.G.") or prestacion.startswith("Ecocardiograma"):
                    ws_pdc_img_ecg.append(row)  # Añadir la fila completa

    # Agregar subtotal en la hoja PDC_IMG_ECG
    if 'PDC_IMG_ECG' in wb.sheetnames:
        ws_pdc_img_ecg = wb['PDC_IMG_ECG']

        # Identificar las columnas TOTAL, CLP y UTILIDAD
        col_total = col_clp = col_utilidad = None
        for col in ws_pdc_img_ecg.iter_cols(1, ws_pdc_img_ecg.max_column):
            header = col[0].value
            if header == 'TOTAL': col_total = col[0].column_letter
            elif header == 'CLP': col_clp = col[0].column_letter
            elif header == 'UTILIDAD': col_utilidad = col[0].column_letter

        # Calcular los subtotales
        if col_total and col_clp and col_utilidad:
            total_sum = clp_sum = utilidad_sum = 0
            for row in range(2, ws_pdc_img_ecg.max_row + 1):
                total_val = ws_pdc_img_ecg[f'{col_total}{row}'].value or 0
                clp_val = ws_pdc_img_ecg[f'{col_clp}{row}'].value or 0
                utilidad_val = ws_pdc_img_ecg[f'{col_utilidad}{row}'].value or 0

                total_sum += total_val
                clp_sum += clp_val
                utilidad_sum += utilidad_val

            # Agregar una fila de subtotal
            ws_pdc_img_ecg.append([])  # Añadir una fila vacía
            subtotal_row = ws_pdc_img_ecg.max_row + 1
            ws_pdc_img_ecg[f'A{subtotal_row}'] = "SUBTOTAL"
            ws_pdc_img_ecg[f'{col_total}{subtotal_row}'] = round(total_sum)
            ws_pdc_img_ecg[f'{col_clp}{subtotal_row}'] = round(clp_sum)
            ws_pdc_img_ecg[f'{col_utilidad}{subtotal_row}'] = round(utilidad_sum)

    # Obtener el valor del subtotal de CLP de la hoja PDC_IMG_ECG
    if 'PDC_IMG_ECG' in wb.sheetnames:
        ws_pdc_img_ecg = wb['PDC_IMG_ECG']
        
        # Buscar la fila que contiene "SUBTOTAL" en la columna A
        subtotal_clp_value = None
        for row in ws_pdc_img_ecg.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] == "SUBTOTAL":
                # Obtener el valor de la columna CLP (columna S) en la misma fila
                subtotal_clp_value = ws_pdc_img_ecg[f'S{ws_pdc_img_ecg.max_row}'].value
                break

        # Si se encontró el valor, llevarlo a la hoja 8-Rentabilidad de la Unidad en C5
        if subtotal_clp_value is not None:
            ws_rentabilidad = wb['8-Rentabilidad de la Unidad']
            ws_rentabilidad['C5'] = subtotal_clp_value
            
            # Calcular la UTILIDAD como C5 - B5
            if ws_rentabilidad['B5'].value is not None:
                ws_rentabilidad['D5'] = ws_rentabilidad['C5'].value - ws_rentabilidad['B5'].value
            else:
                print("Advertencia: No se pudo calcular la UTILIDAD porque B5 está vacío.")
        else:
            print("Advertencia: No se encontró el subtotal de CLP en la hoja PDC_IMG_ECG.")
    else:
        print("Advertencia: No se encontró la hoja PDC_IMG_ECG.")

    # Crear la hoja PDC_IMG_NONECG con la información filtrada de PDC_IMG
    if 'PDC_IMG' in wb.sheetnames:
        ws_pdc_img = wb['PDC_IMG']
        ws_pdc_img_nonecg = wb.create_sheet(title='PDC_IMG_NONECG')

        # Copiar encabezados como una fila
        headers = [cell.value for cell in ws_pdc_img[1]]
        ws_pdc_img_nonecg.append(headers)  # Añadir los encabezados como una fila

        # Encontrar el índice de la columna 'PRESTACION'
        prestacion_col_idx = None
        for idx, cell in enumerate(ws_pdc_img[1]):
            if cell.value == 'PRESTACION':
                prestacion_col_idx = idx
                break

        if prestacion_col_idx is None:
            print("No se encontró la columna 'PRESTACION' en la hoja PDC_IMG.")
        else:
            # Filtrar filas y copiar datos
            for row in ws_pdc_img.iter_rows(min_row=2, values_only=True):
                prestacion = str(row[prestacion_col_idx])
                if not (prestacion.startswith("E.C.G.") or prestacion.startswith("Ecocardiograma")):
                    ws_pdc_img_nonecg.append(row)  # Añadir la fila completa

    # Agregar subtotal en la hoja PDC_IMG_NONECG
    if 'PDC_IMG_NONECG' in wb.sheetnames:
        ws_pdc_img_nonecg = wb['PDC_IMG_NONECG']

        # Identificar las columnas TOTAL, CLP y UTILIDAD
        col_total = col_clp = col_utilidad = None
        for col in ws_pdc_img_nonecg.iter_cols(1, ws_pdc_img_nonecg.max_column):
            header = col[0].value
            if header == 'TOTAL': col_total = col[0].column_letter
            elif header == 'CLP': col_clp = col[0].column_letter
            elif header == 'UTILIDAD': col_utilidad = col[0].column_letter

        # Calcular los subtotales
        if col_total and col_clp and col_utilidad:
            total_sum = clp_sum = utilidad_sum = 0
            for row in range(2, ws_pdc_img_nonecg.max_row + 1):
                total_val = ws_pdc_img_nonecg[f'{col_total}{row}'].value or 0
                clp_val = ws_pdc_img_nonecg[f'{col_clp}{row}'].value or 0
                utilidad_val = ws_pdc_img_nonecg[f'{col_utilidad}{row}'].value or 0

                total_sum += total_val
                clp_sum += clp_val
                utilidad_sum += utilidad_val

            # Agregar una fila de subtotal
            ws_pdc_img_nonecg.append([])  # Añadir una fila vacía
            subtotal_row = ws_pdc_img_nonecg.max_row + 1
            ws_pdc_img_nonecg[f'A{subtotal_row}'] = "SUBTOTAL"
            ws_pdc_img_nonecg[f'{col_total}{subtotal_row}'] = round(total_sum)
            ws_pdc_img_nonecg[f'{col_clp}{subtotal_row}'] = round(clp_sum)
            ws_pdc_img_nonecg[f'{col_utilidad}{subtotal_row}'] = round(utilidad_sum)

        # Obtener el valor del subtotal de CLP de la hoja PDC_IMG_NONECG
        subtotal_clp_value = ws_pdc_img_nonecg[f'{col_clp}{subtotal_row}'].value

        # Si se encontró el valor, llevarlo a la hoja 8-Rentabilidad de la Unidad en C3
        if subtotal_clp_value is not None:
            ws_rentabilidad = wb['8-Rentabilidad de la Unidad']
            ws_rentabilidad['C3'] = subtotal_clp_value
            
            # Calcular la UTILIDAD como C3 - B3
            if ws_rentabilidad['B3'].value is not None:
                ws_rentabilidad['D3'] = ws_rentabilidad['C3'].value - ws_rentabilidad['B3'].value
            else:
                print("Advertencia: No se pudo calcular la UTILIDAD porque B3 está vacío.")
        else:
            print("Advertencia: No se encontró el subtotal de CLP en la hoja PDC_IMG_NONECG.")
    else:
        print("Advertencia: No se encontró la hoja PDC_IMG_NONECG.")

    # Contar celdas vacías en la columna S de la hoja PDC_IMG_NONECG
    if 'PDC_IMG_NONECG' in wb.sheetnames:
        ws_pdc_img_nonecg = wb['PDC_IMG_NONECG']
        ws_rentabilidad['E3'] = count_empty_cells(ws_pdc_img_nonecg, 'S')  # Contar celdas vacías en la columna S de PDC_IMG_NONECG

    # Contar celdas vacías en la columna S de la hoja PDC_IMG_ECG
    if 'PDC_IMG_ECG' in wb.sheetnames:
        ws_pdc_img_ecg = wb['PDC_IMG_ECG']
        ws_rentabilidad['E5'] = count_empty_cells(ws_pdc_img_ecg, 'S')  # Contar celdas vacías en la columna S de PDC_IMG_ECG

    # Mover la hoja "8-Rentabilidad de la Unidad" al final
    wb._sheets = [sheet for sheet in wb._sheets if sheet.title != "8-Rentabilidad de la Unidad"] + [wb['8-Rentabilidad de la Unidad']]

    # Guardar el archivo
    wb.save(output_path)

    print(f"Archivo generado en: {output_path}")

if __name__ == "__main__":
    main()